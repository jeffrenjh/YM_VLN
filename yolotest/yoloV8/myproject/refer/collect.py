import sys
import socket
import threading
import json
from RealMan import RM_controller
from tf.transformations import euler_from_quaternion
from Robotic_Arm.rm_robot_interface import *
import os
import cv2
import time
import math
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.spatial.transform import Rotation as R  # 需要安装 scipy
from data_dual import CollectData
from collect_data import calculate_movement_delta, should_save_frame, initialize_realsense, initialize_cameras  
from queue import Queue
import pyrealsense2 as rs

DEBUG = False

def euler_to_quat(rx, ry, rz):
    # 欧拉角转四元数，单位为度
    r = R.from_euler('xyz', [rx, ry, rz], degrees=True)
    q = r.as_quat()  # [x, y, z, w]
    return {"x": q[0], "y": q[1], "z": q[2], "w": q[3]}

def debug_print(msg, release=False):
    if release or DEBUG:
        print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] [DEBUG] {msg}")

msg_freq = 10

# 初始化机械臂控制器
left_wrist_controller = RM_controller("192.168.0.18", rm_thread_mode_e.RM_TRIPLE_MODE_E)
right_wrist_controller = RM_controller("192.168.0.19")

# import pdb
# pdb.set_trace()
# 设置机械臂初始状态
left_wrist_controller.arm_controller.rm_set_gripper_release(500, True, 1)
right_wrist_controller.arm_controller.rm_set_gripper_release(500, True, 1)

debug_print("机械臂初始化完成", True)

# Socket 配置
HOST = '192.168.0.20'  # 替换为服务器 IP 地址
PORT = 12345            # 替换为服务器端口号

collect_index = 0
output_path = None
last_left_arm_union = None
last_right_arm_union = None
last_left_gripper = 0.0
last_right_gripper = 0.0
# 相机管道（若有，需初始化）
pipelines = {}  # 示例：实际使用时替换为相机初始化后的管道

save_queue = Queue()
stop = False
def save_worker():
    global stop
    # print("1")
    while True:
        print("[Worker] 等待保存路径...")
        print(f"save_queue's length: {save_queue.qsize()}")
        # print(output_path)
        output_path = save_queue.get()
        if stop==True:

            break
        if output_path is None:
            break
        try:
            debug_print(f"保存数据到------------- {output_path}")
            collect_and_save(output_path)
            time.sleep(0.05)
        except Exception as e:
            print(f"[Worker ERROR] {e}")
        finally:
            save_queue.task_done()

save_thread = threading.Thread(target=save_worker, daemon=True)
save_thread.start()
print("[DEBUG] save_worker 启动")

# def save_collect_data(right_arm_union, right_gripper, left_arm_union, left_gripper, imgs, output_path, index):
#     print("[DEBUG] save_collect_data() 函数被调用")
#     os.makedirs(output_path, exist_ok=True)
#     file_path = os.path.join(output_path, f"{index:06d}.npy")

#     try:
#         joint = np.concatenate([right_arm_union["joint"], left_arm_union["joint"]])
#         pos = np.concatenate([right_arm_union["pose"], left_arm_union["pose"]])
#         gripper = np.array([right_gripper, left_gripper])
#         data_dict = {
#             'joint': np.array(joint, dtype=np.float32),  
#             'pose': np.array(pos, dtype=np.float32),    
#             'gripper': np.array(gripper, dtype=np.float32), 
#             'imgs': imgs 
#         }
#         np.save(file_path, data_dict)
#         for camera_name, camera_data in imgs.items():
#             color_img = camera_data['color']
#             color_path = os.path.join(output_path, f"{camera_name}_color_{index:06d}.jpg")
#             cv2.imwrite(color_path, color_img)
#             print(f"[INFO] 彩色图像已保存到 {color_path}")

#         print(f"[INFO] 数据已保存到 {file_path}")
#         print("即将保存到：", os.path.abspath(file_path))
#     except Exception as e:
#         print(f"[ERROR] 保存数据失败: {e}")


# def reset_all_cameras(wait_seconds=5):
#     """重置所有 RealSense 相机并等待设备重新枚举"""
#     ctx = rs.context()
#     devices = ctx.query_devices()
    
#     if not devices:
#         print("未检测到任何 RealSense 设备")
#         return
    
#     for dev in devices:
#         sn = dev.get_info(rs.camera_info.serial_number)
#         print(f"Resetting camera SN: {sn}")
#         try:
#             dev.hardware_reset()
#         except Exception as e:
#             print(f"Warning: 重置相机 {sn} 失败: {e}")
    
#     print(f"等待 {wait_seconds} 秒让相机重新枚举...")
#     time.sleep(wait_seconds)

# # -----------------------------
# # 在 collect.py 的入口处调用
# reset_all_cameras(wait_seconds=5)

START_POSITION_ANGLE_LEFT_ARM = [
    85,     # Joint 1
    -54,     # Joint 2
    -6,    # Joint 3
    -65,     # Joint 4
    -29,      # Joint 5
    -85,    # Joint 6
    80      # Joint 7
]

# Define start position (in degrees)
START_POSITION_ANGLE_RIGHT_ARM = [
    -74,    # Joint 1
    41,     # Joint 2
    -4,   # Joint 3
    84,     # Joint 4
    46,    # Joint 5
    74,   # Joint 6
    103      # Joint 7
]
def initialize_realman():
    """
    初始化 RealMan 机械臂
    """
    global left_wrist_controller, right_wrist_controller

    # 设置左臂初始位置
    left_signal = left_wrist_controller.move_init(START_POSITION_ANGLE_LEFT_ARM)
    debug_print(f"左臂初始位置设置为: {START_POSITION_ANGLE_LEFT_ARM}", True)

    # 设置右臂初始位置
    right_signal = right_wrist_controller.move_init(START_POSITION_ANGLE_RIGHT_ARM)
    debug_print(f"右臂初始位置设置为: {START_POSITION_ANGLE_RIGHT_ARM}", True)

    if left_signal != 0 and right_signal != 0:
        debug_print(f"机械臂初始位置设置失败---rightarm:{right_signal}---leftarm:{left_signal}，请手动示教机械臂", True)
        
        return False

    # 等待机械臂到达初始位置
    time.sleep(2)
    debug_print("机械臂初始位置设置成功", True)
    return True
def save_collect_data(right_arm_union, right_gripper, left_arm_union, left_gripper, imgs, output_path, index,
                      L_img_ts, L_pose_ts, R_img_ts, R_pose_ts, C_img_ts):
    """
    保存采集数据到 Excel + 图片（防文件损坏版）
    """
    print("[DEBUG] save_collect_data() 被调用")
    os.makedirs(output_path, exist_ok=True)

    # ---------------- 1. 处理数据 ----------------
    seq_num = index
    L_joint = list(left_arm_union["joint"])
    L_pose = list(left_arm_union["pose"])
    L_grip = left_gripper
    R_joint = list(right_arm_union["joint"])
    R_pose = list(right_arm_union["pose"])
    R_grip = right_gripper
    time_stamps = [L_img_ts, L_pose_ts, R_img_ts, R_pose_ts, C_img_ts]

    row_data = [seq_num] + L_joint + L_pose + [L_grip] + R_joint + R_pose + [R_grip] + time_stamps

    excel_path = os.path.join(output_path, "collect_data.xlsx")

    headers = (
        ["序号"] +
        [f"L_joint_{i+1}" for i in range(7)] +
        [f"L_pose_{i+1}" for i in range(6)] +
        ["L_gripper"] +
        [f"R_joint_{i+1}" for i in range(7)] +
        [f"R_pose_{i+1}" for i in range(6)] +
        ["R_gripper"] +
        ["L_img_ts", "L_pose_ts", "R_img_ts", "R_pose_ts", "C_img_ts"]
    )

    # ---------------- 2. 写入 Excel ----------------
    # 检查文件是否损坏
    if not os.path.exists(excel_path):
        # 文件不存在，创建并写表头
        df = pd.DataFrame([row_data], columns=headers)
        df.to_excel(excel_path, index=False)
        print(f"[INFO] Excel 文件不存在，已创建并写入: {excel_path}")
    else:
        # 文件存在，追加数据
        book = load_workbook(excel_path)
        start_row = book['Sheet1'].max_row
        book.close()

        df = pd.DataFrame([row_data])
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, index=False, header=False, startrow=start_row)
        print(f"[INFO] Excel 文件已存在，追加一行数据: {excel_path}")

    # ---------------- 3. 保存图片 ----------------
    # img_map = {
    #     1: imgs.get("left", {}).get("color"),
    #     2: imgs.get("right", {}).get("color"),
    #     3: imgs.get("center", {}).get("color")
    # }
    camera_id_map = {
        "head": 1,
        "left_wrist": 2,
        "right_wrist": 3
    }

    # for cam_id, img in img_map.items():
    #     if img is not None:
    #         img_path = os.path.join(output_path, f"{seq_num}_{cam_id}.jpg")
    #         cv2.imwrite(img_path, img)
    #         print(f"[INFO] 图片已保存 {img_path}")
    #     else:
    #         print(f"[WARN] 相机 {cam_id} 图像为空，未保存。")

    for camera_name, camera_data in imgs.items():
        color_img = camera_data['color']
        cam_id = camera_id_map.get(camera_name, 0)  # 如果找不到映射就给 0
        print(f"[DEBUG] 相机 {camera_name} 的 ID 是 {cam_id}")
        if cam_id == 0:
            print(f"[WARN] 未知相机名称: {camera_name}")
            continue
        color_path = os.path.join(output_path, f"{seq_num}_{cam_id}.jpg")
        cv2.imwrite(color_path, color_img)
        print(f"[INFO] 彩色图像已保存到 {color_path}")

    print("[INFO] 完成一次数据保存\n")

def collect_and_save(output_path):
    #"""独立函数：收集数据并根据运动过滤决定是否保存"""
    global left_wrist_controller, right_wrist_controller, collect_index
    global last_left_arm_union, last_right_arm_union, last_left_gripper, last_right_gripper, pipelines
    # os.makedirs(output_path, exist_ok=True)

    # print("1"*50)

    try:
        # print("2"*50)
        # 1. 获取当前双臂状态（关节+姿态）
        right_pose_ts = time.time()
        _, right_arm_union = right_wrist_controller.arm_controller.rm_get_current_arm_state()
        left_pose_ts = time.time()
        _, left_arm_union = left_wrist_controller.arm_controller.rm_get_current_arm_state()

        # 2. 获取夹爪状态
        _, right_gripper_data = right_wrist_controller.arm_controller.rm_get_gripper_state()
        right_gripper = float(right_gripper_data['actpos']) / 1000
        _, left_gripper_data = left_wrist_controller.arm_controller.rm_get_gripper_state()
        left_gripper = float(left_gripper_data['actpos']) / 1000

        # 3. 获取图像数据（复用之前的collect函数逻辑）
        imgs = {}
        timestamps = {}
        # print("3"*50)
        for camera_name in pipelines.keys():
            print(f"[DEBUG] 获取相机 {camera_name} 的数据")
            frames = pipelines[camera_name].wait_for_frames()
            color_frame = frames.get_color_frame()

            # depth_frame = frames.get_depth_frame()
            if color_frame: #and depth_frame:
                # 显示当前帧
                color_image = np.asanyarray(color_frame.get_data())
                cv2.imshow(f'{camera_name} Camera', color_image)
                cv2.waitKey(1)

                color_ts = color_frame.get_timestamp() / 1000.0
                imgs[camera_name] = {
                    'color': np.asanyarray(color_frame.get_data()),
                    # 'depth': np.asanyarray(depth_frame.get_data())
                }
                timestamps[camera_name] = color_ts
            else:
                debug_print(f"相机 {camera_name} 数据获取失败", True)
                return  # 图像获取失败则不保存
        # print("4.1"*50)
        # 4. 用已有CollectData类整合数据
        # collected_data = CollectData(
        #     right_arm_union=right_arm_union,
        #     right_gripper=right_gripper,
        #     left_arm_union=left_arm_union,
        #     left_gripper=left_gripper,
        #     imgs=imgs
        # )

        vis_img = np.zeros((400, 800, 3), dtype=np.uint8)
        # 显示的文字内容
        # print("4"*50)
        if imgs:  # 如果至少有一个相机
            main_camera = list(imgs.keys())[1]  
            color_img = imgs[main_camera]['color'].copy()
            # print("5"*50)
            # 显示文字信息
            text_lines = [
                f"[INFO VISUALIZATION]",
                f"Collect Index: {collect_index}",
                f"Output Path: {output_path}",
                f"Right Gripper: {right_gripper:.3f}",
                f"Left Gripper: {left_gripper:.3f}",
                f"last_left_arm_union: {last_left_arm_union}",
                f"last_right_arm_union: {last_right_arm_union}",
            ]

            for i, line in enumerate(text_lines):
                cv2.putText(color_img, line, (10, 30 + i * 25), cv2.FONT_HERSHEY_SIMPLEX,
                            0.6, (0, 255, 0), 2)
        text_lines.append(f"Right Gripper Raw: {right_gripper_data['actpos']}")
        text_lines.append(f"Left Gripper Raw: {left_gripper_data['actpos']}")

        # 5. 运动过滤（复用之前的should_save_frame）
        if last_left_arm_union is not None and last_right_arm_union is not None:
            # text_lines.append("[Check] Entered save decision logic1 ") 
            # 计算关节运动差异
            left_delta = calculate_movement_delta(left_arm_union["joint"], last_left_arm_union["joint"])
            right_delta = calculate_movement_delta(right_arm_union["joint"], last_right_arm_union["joint"])
            delta = np.concatenate([left_delta, right_delta])

            # 计算夹爪变化
            left_gripper_change = left_gripper - last_left_gripper
            right_gripper_change = right_gripper - last_right_gripper

            # 判断是否保存
            # text_lines.append("[Check] Entered save decision logic2 ") 
            # save_collect_data(
            #     right_arm_union, right_gripper,
            #     left_arm_union, left_gripper,
            #     imgs, output_path, collect_index
            # )
            text_lines.append("[Check] Entered save decision logic5 ") 
            if should_save_frame(delta, left_gripper_change, right_gripper_change):
                text_lines.append("[Check] Entered save decision logic3 ") 
                # collected_data.write(output_path, collect_index)  # 保存为npy
                # save_collect_data(
                #     right_arm_union, right_gripper,
                #     left_arm_union, left_gripper,
                #     imgs, output_path, collect_index
                # )
                L_img_ts=timestamps.get("head")
                L_pose_ts=left_pose_ts   # 机械臂位姿时间戳
                R_img_ts=timestamps.get("left_wrist")
                R_pose_ts=right_pose_ts  # 机械臂位姿时间戳
                C_img_ts=timestamps.get("right_wrist")
                # print("L_img_ts:", L_img_ts)
                # print("L_pose_ts:", L_pose_ts)
                # print("R_img_ts:", R_img_ts)
                # print("R_pose_ts:", R_pose_ts)
                # print("时间戳:", [L_img_ts, L_pose_ts, R_img_ts, R_pose_ts, C_img_ts])
                save_collect_data(
                    right_arm_union, right_gripper,
                    left_arm_union, left_gripper,
                    imgs, output_path, collect_index,
                    L_img_ts, L_pose_ts,   # 机械臂位姿时间戳
                    R_img_ts, R_pose_ts, C_img_ts
                )
                text_lines.append("[Check] Entered save decision logic4 ") 
                debug_print(f"保存帧 {collect_index} 到 {output_path}", True)
                # global collect_index
                collect_index += 1

        # 加入左右臂关节数据（取前3个关节看就够了）
        if right_arm_union is not None:
            text_lines.append(f"Right Arm Joint[0:3]: {np.round(right_arm_union['joint'][:3], 2)}")
        if left_arm_union is not None:
            text_lines.append(f"Left Arm Joint[0:3]: {np.round(left_arm_union['joint'][:3], 2)}")

        # 把文字写到图像上
        for i, line in enumerate(text_lines):
            cv2.putText(vis_img, line, (20, 30 + i * 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

        # 显示窗口
        cv2.imshow("Data Visualization", vis_img)
        cv2.waitKey(1)

        # 6. 更新上一帧状态
        last_left_arm_union = left_arm_union
        last_right_arm_union = right_arm_union
        last_left_gripper = left_gripper
        last_right_gripper = right_gripper

    except Exception as e:
        debug_print(f"collect_and_save 出错: {e}", True)

def handle_socket_data(data):
    """
    处理从 socket 接收到的数据
    """
    global right_wrist_controller, left_wrist_controller, collect_index, output_path
    global last_left_arm_union, last_right_arm_union, last_left_gripper, last_right_gripper, pipelines
    global stop
    try:
        # 将 JSON 格式的数据解析为字典
        data_dict = json.loads(data)
        # print(data_dict)

        if data_dict['buttonA']:
            stop = True
            

        # 提取左臂位置和旋转角度
        left_pos = data_dict['leftPos']
        left_rot = data_dict['leftRot']
        left_quat = data_dict['leftQuat']
        x_l, y_l, z_l = left_pos['x'], left_pos['y'], left_pos['z']
        # roll_l, pitch_l, yaw_l = left_rot['x']*math.pi/180, left_rot['y']*math.pi/180, left_rot['z']*math.pi/180
        quat_l = [left_quat['x'], left_quat['y'], left_quat['z'], left_quat['w']]
        roll_l,pitch_l, yaw_l = euler_from_quaternion(quat_l)
        # roll_l = -roll_l  # 翻转 pitch 角度

        # 提取右臂位置和旋转角度
        right_pos = data_dict['rightPos']
        right_rot = data_dict['rightRot']
        right_quat = data_dict['rightQuat']
        x_r, y_r, z_r = right_pos['x'], right_pos['y'], right_pos['z']
        # roll_r, pitch_r, yaw_r = right_rot['x']*math.pi/180, right_rot['y']*math.pi/180, right_rot['z']*math.pi/180
        quat_r = [right_quat['x'], right_quat['y'], right_quat['z'], right_quat['w']]
        roll_r,pitch_r, yaw_r = euler_from_quaternion(quat_r)
        # roll_r = -roll_r  # 翻转 pitch 角度

        # 提取抓手状态
        left_trigger = 1 - data_dict['leftTrigger']
        right_trigger = 1 - data_dict['rightTrigger']

        if (x_l == 0 and y_l == 0 and z_l == 0) : # the position missing, discared
            debug_print("左手坐标为0，丢弃该条信息", True)
        else:

            if left_wrist_controller.is_controlling is False:
                if data_dict['leftGrip']==True:
                    left_wrist_controller.is_controlling = True
                    left_wrist_controller.arm_first_state = left_wrist_controller.get_state()
                    left_wrist_controller.prev_tech_state = [x_l, y_l, z_l, roll_l, pitch_l, yaw_l]
            else:
                if data_dict['leftGrip']==False:
                    left_wrist_controller.is_controlling = False
                    left_wrist_controller.arm_first_state = None
                    left_wrist_controller.prev_tech_state = None
                    debug_print("左臂控制已停止", True)
                    return
                else:
                    # 控制左臂
                    left_wrist_controller.move([x_l, y_l, z_l, roll_l, pitch_l, yaw_l])
                    left_wrist_controller.set_gripper(left_trigger)
                    # debug_print(f"左臂位置: {x_l}, {y_l}, {z_l} | 旋转: {roll_l}, {pitch_l}, {yaw_l}", True)
                    # debug_print(f"左抓手: {left_trigger}", True)

        
        if x_r == 0 and y_r == 0 and z_r == 0:
            debug_print("右手坐标为0，丢弃该条信息", True)
        else:
            if right_wrist_controller.is_controlling is False:
                if data_dict['rightGrip']==True:
                    right_wrist_controller.is_controlling = True
                    right_wrist_controller.arm_first_state = right_wrist_controller.get_state()
                    right_wrist_controller.prev_tech_state = [x_r, y_r, z_r, roll_r, pitch_r, yaw_r]
            else:
                if data_dict['rightGrip']==False:
                    right_wrist_controller.is_controlling = False
                    right_wrist_controller.arm_first_state = None
                    right_wrist_controller.prev_tech_state = None
                    debug_print("右臂控制已停止", True)
                    return
                else:
                    # 控制右臂
                    right_wrist_controller.move([x_r, y_r, z_r, roll_r, pitch_r, yaw_r])
                    right_wrist_controller.set_gripper(right_trigger)
                    # debug_print(f"右臂位置: {x_r}, {y_r}, {z_r} | 旋转: {roll_r}, {pitch_r}, {yaw_r}", True)
                    # debug_print(f"右抓手: {right_trigger}", True)

        # import pdb
        # pdb.set_trace()
        # print(f"将保存数据路径存入queue{output_path}")
        # print(f"save_queue的内容队列长度为{save_queue.qsize()}")
        save_queue.put((output_path))
        # time.sleep(0.3)    # 0.05s 
    except Exception as e:
        debug_print(f"处理数据时出错: {e}", True)

def feedback_thread_func(sock,delay = 0.01):
    global left_wrist_controller,right_wrist_controller
    while True:
        # 读取当前位置并反馈
        left_state = left_wrist_controller.get_state()
        right_state = right_wrist_controller.get_state()
        # print(left_state)
        # print(left_wrist_controller.arm_controller.rm_get_current_arm_state())
        # pos = robot.read_act_pos()
        left_quat = euler_to_quat(left_state[3], left_state[4], left_state[5])
        right_quat = euler_to_quat(right_state[3], right_state[4], right_state[5])
        feedback = {
            "leftPos":{
                "x": left_state[0],
                "y": left_state[1],
                "z": left_state[2],
            },
            "leftRot":{
                "x": left_state[3]*180/math.pi,
                "y": left_state[4]*180/math.pi,
                "z": left_state[5]*180/math.pi,
            },
            "leftQuat": left_quat,
            "rightPos":{
                "x": right_state[0],
                "y": right_state[1],
                "z": right_state[2],
            },
            "rightRot":{
                "x": right_state[3]*180/math.pi,
                "y": right_state[4]*180/math.pi,
                "z": right_state[5]*180/math.pi,
            },
            "rightQuat": right_quat
        }
        feedback_json = json.dumps(feedback) + "\n"
        sock.sendall(feedback_json.encode("utf-8"))
        time.sleep(delay)

def socket_listener():
    """
    Socket 客户端，用于接收位置和抓手信息
    """
    global stop
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.connect((HOST, PORT))
        print(f"[连接成功] 已连接到 Unity 服务器 {HOST}:{PORT}")

        # 启动反馈线程
        # feedback_thread = threading.Thread(target=feedback_thread_func, args=(s,), daemon=True)
        # feedback_thread.start()
        
        buffer = ""
        while True:
            try:
                data = s.recv(1024)
                if not data:
                    print("[断开连接]")
                    break
                buffer += data.decode('utf-8')
                while '\n' in buffer:
                    line, buffer = buffer.split('\n', 1)
                    if line.strip() == "":
                        continue
                    try:
                        msg = json.loads(line)
                        # print("[收到数据]", msg)
                        # debug_print(f"[收到数据] {msg}")
                        # print(f"[收到数据] {msg}")
                        import pdb
                        # pdb.set_trace()
                        handle_socket_data(line)
                        if stop==True:
                            print("结束采集")
                            s.close()
                            cv2.destroyAllWindows()
                            return
                    except json.JSONDecodeError as e:
                        print("[JSON解析失败]", e)
            except Exception as e:
                debug_print(f"Socket接收异常: {e}", True)
                break

if __name__ == '__main__':
    # 配置输出路径
    arm_brand = "RealMan"
    # task_name = "pick_place-single-left-box" #dualarm
    task_name = "test"
    # episode_index = 3
    # output_path = f"./datasets/npy/{arm_brand}/{task_name}/{episode_index}"
    # if not os.path.exists(output_path):
    #     os.makedirs(output_path)

    base_path = f"./datasets/npy/{arm_brand}/{task_name}"
    if not os.path.exists(base_path):
        os.makedirs(base_path)
        episode_index = 0
    else:
        # 获取所有子文件夹名，并筛选出纯数字的
        subfolders = [f for f in os.listdir(base_path) if f.isdigit()]
        if subfolders:
            # 找到最大数字 + 1
            episode_index = max(map(int, subfolders)) + 1
        else:
            episode_index = 0

    output_path = os.path.join(base_path, str(episode_index))
    os.makedirs(output_path, exist_ok=True)

    debug_print(f"数据将保存到: {output_path}", True)
    ctx, devices = initialize_realsense()  # 调用你的函数获取设备列表
    pipelines = initialize_cameras(devices)
    time.sleep(5)
    # print("Pipelines initialized:", pipelines.keys())

    ini_signal = initialize_realman()
    if not ini_signal:
        exit(0)

    # 启动 Socket 监听线程
    thread = threading.Thread(target=socket_listener, daemon=True)
    thread.start()
    thread.join()
    # socket_listener()  # 启动 Socket 监听
    # cv2.destroyAllWindows()


